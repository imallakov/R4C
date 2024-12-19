from django.db.models import Count
from django.http import JsonResponse, HttpResponse
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.http import require_http_methods
from .models import Robot
import json
from openpyxl import Workbook
from datetime import datetime, timedelta


class RobotView(View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            model = data.get('model')
            version = data.get('version')
            created = data.get('created')

            if not Robot.objects.filter(model=model).exists():
                return JsonResponse({"error": "Такой модели роботов не существует."}, status=400)

            serial = f"{model}-{version}"
            robot = Robot(serial=serial, model=model, version=version, created=created)
            robot.save()

            return JsonResponse({"serial": robot.serial, "model": robot.model, "version": robot.version}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Неверный формат JSON."}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


@require_http_methods(["GET"])
def download_robot_production_report(request):
    wb = Workbook()

    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=7)

    robots_last_week = Robot.objects.filter(created__range=(start_date, end_date))

    models = Robot.objects.values_list('model', flat=True).distinct()

    for model in models:
        ws = wb.create_sheet(title=model)

        headers = ['Модель', 'Версия', 'Количество за неделю']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        model_robots = robots_last_week.filter(model=model)

        versions = model_robots.values('version').annotate(count=Count('id'))

        row_num = 2
        for version in versions:
            ws.cell(row=row_num, column=1, value=model)
            ws.cell(row=row_num, column=2, value=version['version'])
            ws.cell(row=row_num, column=3, value=version['count'])
            row_num += 1

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="RobotProductionSummary.xlsx"'
    wb.save(response)

    return response
