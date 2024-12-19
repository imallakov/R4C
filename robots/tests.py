from django.test import TestCase, Client
from django.urls import reverse
import json
from .models import Robot
from django.utils import timezone
from io import BytesIO
from openpyxl import load_workbook


class RobotViewTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        Robot.objects.create(serial="R2-D2", model="R2", version="D2", created="2022-12-31 23:59:59")

    def test_add_robot_post(self):
        url = reverse('add_robot')
        data = {"model": "R2", "version": "D2", "created": "2023-01-01 00:00:00"}
        response = self.client.post(url, data=json.dumps(data), content_type='application/json')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(Robot.objects.count(), 2)

    def test_add_robot_invalid_data(self):
        url = reverse('add_robot')
        data = {"model": "R8", "version": "D2", "created": "2023-01-01 00:00:00"}
        response = self.client.post(url, data=json.dumps(data), content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertIn("Такой модели роботов не существует.", response.json().get('error'))

    def test_method_not_allowed(self):
        url = reverse('add_robot')
        response = self.client.put(url, {}, content_type='application/json')
        self.assertEqual(response.status_code, 405)


class DownloadRobotProductionReportTest(TestCase):
    def setUp(self):
        # Create some test data
        now = timezone.now()
        Robot.objects.create(serial="R0001", model="R2", version="D2", created=now - timezone.timedelta(days=1))
        Robot.objects.create(serial="R0002", model="R2", version="A1", created=now - timezone.timedelta(days=2))
        Robot.objects.create(serial="R0003", model="R2", version="D2", created=now - timezone.timedelta(days=3))
        Robot.objects.create(serial="R0004", model="R3", version="X1", created=now - timezone.timedelta(days=4))

    def test_download_report(self):
        url = reverse('download_report')
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertEqual(response['Content-Disposition'], 'attachment; filename="RobotProductionSummary.xlsx"')

        wb = load_workbook(BytesIO(response.content))

        self.assertIn('R2', wb.sheetnames)
        self.assertIn('R3', wb.sheetnames)

        ws_r2 = wb['R2']
        self.assertEqual(ws_r2['A1'].value, 'Модель')
        self.assertEqual(ws_r2['B1'].value, 'Версия')
        self.assertEqual(ws_r2['C1'].value, 'Количество за неделю')

        data_rows = list(ws_r2.iter_rows(min_row=2, values_only=True))
        self.assertIn(('R2', 'D2', 2), data_rows)
        self.assertIn(('R2', 'A1', 1), data_rows)

        ws_r3 = wb['R3']
        self.assertEqual(ws_r3['A1'].value, 'Модель')
        self.assertEqual(ws_r3['B1'].value, 'Версия')
        self.assertEqual(ws_r3['C1'].value, 'Количество за неделю')
        data_rows_r3 = list(ws_r3.iter_rows(min_row=2, values_only=True))
        self.assertIn(('R3', 'X1', 1), data_rows_r3)
